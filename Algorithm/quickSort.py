import sys, os
from openpyxl import load_workbook

def quick_sort(arr):
    def sort(low, high):
        if high <= low:
            return

        mid = partition(low, high)
        sort(low, mid - 1)
        sort(mid, high)

    def partition(low, high):
        pivot = arr[(low + high) // 2]
        while low <= high:
            while arr[low] < pivot:
                low += 1
            while arr[high] > pivot:
                high -= 1
            if low <= high:
                arr[low], arr[high] = arr[high], arr[low]
                low, high = low + 1, high - 1
        return low

    return sort(0, len(arr) - 1)

path = (os.path.sep.join(sys.argv[0].split(os.path.sep)[:-1]))
final_path = '{}\\input_quick_sort.xlsx'.format(path)
load_wb = load_workbook(final_path, data_only=True)
load_ws = load_wb['Sheet1']

listA = []
for i in range(1,101) :
    listA.append(load_ws.cell(i,1).value)

print("정렬 전 : ", listA)
quick_sort(listA)
print("\n정렬 후 : ", listA)