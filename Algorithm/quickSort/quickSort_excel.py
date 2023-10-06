import sys, os
from openpyxl import load_workbook, Workbook

def partition(arr, start, end):
    pivot = arr[start] # 처음 오는 index를 피봇으로 삼는다.
    i = start # 비교 대상 i와 j
    j = end
    while i<j :
        while (arr[i] <= pivot) & (i < end) : 
            i+=1 # arr[i]가 피봇보다 작거나 같고 index i가 파티션 마지막 index보다 작을 때까지 반복.
        while (arr[j] >= pivot) & (j > start) :
            j-=1 # arr[j]가 피봇보다 크거나 같고 index j가 파티션 마지막 index보다 클 때까지 반복.
        if i < j :
            arr[i], arr[j] = arr[j], arr[i] # 피봇 비교 후 값이 엇갈리는 때엔 i와 j를 서로 교환함.
    arr[start], arr[j] = arr[j], arr[start] # 처음 index와 최종 변경된 end index와 교환해서 고정한다.
    return arr,j # 이후 고정점 j를 기준으로 양 옆의 부분들을 다시 quick_sort한다.

def quick_sort(arr, start, end):
    if start < end:
        arr, return_index = partition(arr, start, end) # 파티션에서 봤 듯 마지막 index의 값을 return해준다.
        quick_sort(arr, start, return_index-1) # 재귀를 이용한 분할 sorting
        quick_sort(arr, return_index+1, end)
 
path = (os.path.sep.join(sys.argv[0].split(os.path.sep)[:-1]))
read_path = '{}\\input_quick_sort.xlsx'.format(path)

load_wb = load_workbook(read_path, data_only=True)
load_ws = load_wb.worksheets[0]

listA = []
for i in range(1,load_ws.max_row + 1) :
    listA.append(load_ws.cell(i,1).value)

print("정렬 전 : ", listA)
quick_sort(listA, 0, len(listA) - 1)
print("\n정렬 후 : ", listA)

write_wb = Workbook()
write_ws = write_wb.active
write_ws.title = 'sheet1'

for i in range(load_ws.max_row) :
    write_ws.cell(row = i+1, column = 1, value = listA[i])

write_wb.save('{}\\output_quick_sort.xlsx'.format(path))